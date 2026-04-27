"""Build an .xlsx export of the AI comparison analysis.

Pure function — takes the parsed `analysis_result` dict produced by
`TenderService.analyse_offers` and returns xlsx bytes. The router
streams the bytes back to the broker as a download.

Layout:
- Sheet "Sammenligning" — one block per category, with the field labels
  in column A, one column per insurer, and a "Kommentar" column at the end
- The AI's recommended insurer header gets a green highlight
- Top of the sheet: the title, the AI recommendation summary, and the
  short oppsummering
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


_HEADER_FILL = PatternFill("solid", fgColor="E5E7EB")
_RECOMMEND_FILL = PatternFill("solid", fgColor="DCFCE7")
_TITLE_FONT = Font(name="Calibri", size=14, bold=True)
_SECTION_FONT = Font(name="Calibri", size=11, bold=True)


def build_comparison_xlsx(analysis: dict[str, Any], tender_title: str) -> bytes:
    """Render the comparison JSON into an .xlsx workbook."""
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sammenligning"

    cursor = _write_header(ws, tender_title, analysis)
    columns = _collect_columns(analysis)
    recommended = (analysis.get("anbefaling") or {}).get("forsikringsgiver")
    cursor = _write_column_header(ws, cursor, columns, recommended)

    for cat in analysis.get("sammenligning") or []:
        cursor = _write_category(ws, cursor, cat, columns)

    _set_column_widths(ws, columns)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_header(ws: Worksheet, title: str, analysis: dict[str, Any]) -> int:
    """Write title + recommendation + summary; return next free row."""
    ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
    rec = analysis.get("anbefaling") or {}
    if rec:
        ws.cell(
            row=2,
            column=1,
            value=f"Anbefaling: {rec.get('forsikringsgiver', '')} — {rec.get('begrunnelse', '')}",
        )
    summary = analysis.get("oppsummering")
    if summary:
        ws.cell(row=3, column=1, value=summary).alignment = Alignment(wrap_text=True)
    return 5


def _collect_columns(analysis: dict[str, Any]) -> list[str]:
    """Walk the comparison and collect insurer names in stable order."""
    names: list[str] = []
    for cat in analysis.get("sammenligning") or []:
        for fld in cat.get("felter") or []:
            for n in (fld.get("verdier") or {}).keys():
                if n not in names:
                    names.append(n)
    return names


def _write_column_header(
    ws: Worksheet, row: int, columns: list[str], recommended: str | None
) -> int:
    """Write the per-insurer column headers; highlight the recommended one."""
    ws.cell(row=row, column=1, value="Felt").font = _SECTION_FONT
    for i, name in enumerate(columns, start=2):
        c = ws.cell(row=row, column=i, value=name)
        c.font = _SECTION_FONT
        c.fill = _RECOMMEND_FILL if name == recommended else _HEADER_FILL
    ws.cell(row=row, column=2 + len(columns), value="Kommentar").font = _SECTION_FONT
    return row + 1


def _write_category(
    ws: Worksheet, row: int, cat: dict[str, Any], columns: list[str]
) -> int:
    """Write a category block (header row + one row per field)."""
    ws.cell(row=row, column=1, value=cat.get("kategori", "")).font = _SECTION_FONT
    row += 1
    for fld in cat.get("felter") or []:
        ws.cell(row=row, column=1, value=fld.get("felt", ""))
        for i, name in enumerate(columns, start=2):
            ws.cell(row=row, column=i, value=(fld.get("verdier") or {}).get(name, ""))
        ws.cell(row=row, column=2 + len(columns), value=fld.get("kommentar", ""))
        row += 1
    return row + 1  # blank spacer between categories


def _set_column_widths(ws: Worksheet, columns: list[str]) -> None:
    ws.column_dimensions[get_column_letter(1)].width = 32
    for i in range(2, 2 + len(columns)):
        ws.column_dimensions[get_column_letter(i)].width = 24
    ws.column_dimensions[get_column_letter(2 + len(columns))].width = 36
