"""Unit tests for api/services/tender_excel.py — comparison .xlsx builder."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from api.services.tender_excel import build_comparison_xlsx


_SAMPLE_ANALYSIS = {
    "anbefaling": {
        "forsikringsgiver": "Gjensidige Forsikring",
        "begrunnelse": "Beste pris/dekning-balanse",
    },
    "oppsummering": "Tre tilbud, Gjensidige sterkest på dekning og pris.",
    "nøkkelforskjeller": ["Egenandel", "Dekning på driftstap"],
    "sammenligning": [
        {
            "kategori": "Premie",
            "felter": [
                {
                    "felt": "Årspremie",
                    "verdier": {
                        "Gjensidige Forsikring": "120 000",
                        "If Skadeforsikring": "135 000",
                        "Tryg Forsikring": "128 000",
                    },
                    "kommentar": "Gjensidige er rimeligst",
                    "konfidens": "høy",
                }
            ],
        },
        {
            "kategori": "Egenandel",
            "felter": [
                {
                    "felt": "Bygning",
                    "verdier": {
                        "Gjensidige Forsikring": "10 000",
                        "If Skadeforsikring": "20 000",
                        "Tryg Forsikring": "15 000",
                    },
                    "kommentar": "",
                    "konfidens": "middels",
                }
            ],
        },
    ],
}


def _load(content: bytes):
    wb = load_workbook(BytesIO(content), data_only=True)
    return wb["Sammenligning"]


def test_returns_non_empty_xlsx_bytes():
    out = build_comparison_xlsx(_SAMPLE_ANALYSIS, "Demo Anbud")
    assert isinstance(out, bytes)
    assert len(out) > 1000  # smallest valid xlsx is ~6kb anyway
    assert out[:2] == b"PK"  # zip magic number — xlsx is a zip


def test_writes_title_recommendation_and_summary():
    ws = _load(build_comparison_xlsx(_SAMPLE_ANALYSIS, "Demo Anbud"))
    assert ws.cell(row=1, column=1).value == "Demo Anbud"
    assert "Gjensidige Forsikring" in str(ws.cell(row=2, column=1).value)
    assert "sterkest" in str(ws.cell(row=3, column=1).value)


def test_columns_are_per_insurer_and_recommended_is_highlighted():
    ws = _load(build_comparison_xlsx(_SAMPLE_ANALYSIS, "Demo Anbud"))
    # Column header row is at row 5; col 1 is "Felt", insurers start at col 2
    assert ws.cell(row=5, column=1).value == "Felt"
    assert ws.cell(row=5, column=2).value == "Gjensidige Forsikring"
    assert ws.cell(row=5, column=3).value == "If Skadeforsikring"
    assert ws.cell(row=5, column=4).value == "Tryg Forsikring"
    # Last column is the comment
    assert ws.cell(row=5, column=5).value == "Kommentar"


def test_renders_each_category_block():
    ws = _load(build_comparison_xlsx(_SAMPLE_ANALYSIS, "Demo Anbud"))
    # Walk down column A and collect non-empty cells
    labels = [
        ws.cell(row=r, column=1).value
        for r in range(1, ws.max_row + 1)
        if ws.cell(row=r, column=1).value
    ]
    assert "Premie" in labels
    assert "Egenandel" in labels
    assert "Årspremie" in labels
    assert "Bygning" in labels


def test_handles_missing_optional_fields_gracefully():
    minimal = {"sammenligning": []}
    out = build_comparison_xlsx(minimal, "Tom analyse")
    ws = _load(out)
    assert ws.cell(row=1, column=1).value == "Tom analyse"


@pytest.mark.parametrize(
    "missing_field",
    [
        {"anbefaling": {}, "sammenligning": []},
        {"oppsummering": "", "sammenligning": []},
        {"nøkkelforskjeller": None, "sammenligning": []},
    ],
)
def test_no_crash_on_partial_analysis(missing_field):
    build_comparison_xlsx(missing_field, "Partial")
